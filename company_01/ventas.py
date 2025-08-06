import pandas as pd
import openpyxl
import xlrd
import os
import csv
from datetime import datetime
from io import BytesIO
import re

def get_column_letter(col_num):
    """Genera letras de columna para Excel (A, B, C, ..., AA, AB, etc.)"""
    letters = []
    for i in range(1, 50):  # Ampliado para cubrir hasta AV (columna 48)
        if i <= 26:
            letters.append(chr(64 + i))
        else:
            first = chr(64 + ((i - 27) // 26) + 1)
            second = chr(64 + ((i - 27) % 26) + 1)
            letters.append(first + second)
    return letters[col_num - 1] if col_num <= len(letters) else f"Col{col_num}"

def clean_value_as_string(value, debug_info=""):
    """Convierte cualquier valor a string limpio, manteniendo el formato original"""
    try:
        if value is None:
            return ""
        elif pd.isna(value):
            return ""
        else:
            # Convertir a string y limpiar espacios
            str_value = str(value).strip()
            
            # Si es 'nan' como string, convertir a vacío
            if str_value.lower() in ['nan', 'none', 'null']:
                return ""
            
            # Si termina en .0, removerlo (para IDs de documentos)
            if str_value.endswith('.0'):
                str_value = str_value[:-2]
                
            return str_value
    except Exception:
        return ""

def clean_numeric_value(value):
    """Convierte un valor numérico a float con 2 decimales"""
    try:
        if value is None or pd.isna(value):
            return 0.00
        
        # Convertir a string y limpiar
        str_value = str(value).strip()
        
        # Si está vacío o es texto no numérico, devolver 0
        if str_value == '' or str_value.lower() in ['nan', 'none', 'null']:
            return 0.00
        
        # Intentar convertir a float
        try:
            float_val = float(str_value)
            return round(float_val, 2)
        except (ValueError, TypeError):
            return 0.00
            
    except (ValueError, TypeError):
        return 0.00

def clean_date_value(value):
    """Convierte fecha del formato dd/mm/aaaa xx:xx:xx al formato dd/mm/aa"""
    try:
        print(f"*** DEBUGGING clean_date INICIO ***")
        print(f"DEBUGGING clean_date: Valor recibido: '{value}'")
        print(f"DEBUGGING clean_date: Tipo de valor: {type(value)}")
        
        if value is None:
            print("DEBUGGING clean_date: Valor es None - RETORNANDO VACÍO")
            return ""
            
        if pd.isna(value):
            print("DEBUGGING clean_date: Valor es NaN - RETORNANDO VACÍO")
            return ""
        
        # Convertir a string
        str_value = str(value).strip()
        print(f"DEBUGGING clean_date: Valor convertido a string: '{str_value}'")
        
        if str_value == '' or str_value.lower() in ['nan', 'none', 'null']:
            print("DEBUGGING clean_date: String vacío o null - RETORNANDO VACÍO")
            return ""
        
        print(f"DEBUGGING clean_date: Buscando patrón dd/mm/aaaa xx:xx:xx...")
        
        # Patrón específico para dd/mm/aaaa xx:xx:xx
        pattern = r'(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{1,2}):(\d{1,2})'
        
        print(f"DEBUGGING clean_date: Aplicando patrón: {pattern}")
        match = re.search(pattern, str_value)
        
        if match:
            print(f"DEBUGGING clean_date: ¡PATRÓN ENCONTRADO! Grupos: {match.groups()}")
            day, month, year, hour, minute, second = match.groups()
            
            print(f"DEBUGGING clean_date: Fecha detectada: día={day}, mes={month}, año={year}")
            print(f"DEBUGGING clean_date: Hora detectada: {hour}:{minute}:{second}")
            
            # Convertir año a 2 dígitos
            year_short = year[-2:]
            print(f"DEBUGGING clean_date: Año convertido a 2 dígitos: {year_short}")
            
            # Formatear con ceros a la izquierda (solo fecha, sin hora)
            result = f"{day.zfill(2)}/{month.zfill(2)}/{year_short}"
            print(f"DEBUGGING clean_date: *** FECHA PROCESADA EXITOSAMENTE: '{result}' ***")
            print(f"*** DEBUGGING clean_date FIN ***")
            return result
        else:
            print(f"DEBUGGING clean_date: *** ERROR - NO SE ENCONTRÓ PATRÓN dd/mm/aaaa xx:xx:xx ***")
            print(f"DEBUGGING clean_date: String analizado: '{str_value}'")
            print(f"*** DEBUGGING clean_date FIN ***")
            return ""
        
    except Exception as e:
        print(f"DEBUGGING clean_date: *** EXCEPCIÓN CAPTURADA: {e} ***")
        print(f"*** DEBUGGING clean_date FIN ***")
        return ""

def is_empty_cell(value):
    """Verifica si una celda está vacía"""
    if value is None:
        return True
        
    if pd.isna(value):
        return True
    
    try:
        str_value = str(value).strip()
        if str_value == '' or str_value.lower() in ['nan', 'none', 'null']:
            return True
        return False
    except Exception:
        return True

def has_content_in_column_e(value):
    """Verifica si hay cualquier contenido en la columna E"""
    try:
        if value is None or pd.isna(value):
            return False
        
        str_value = str(value).strip()
        if str_value == '' or str_value.lower() in ['nan', 'none', 'null']:
            return False
        
        # Si hay cualquier contenido, retornar True
        print(f"🎯 CONTENIDO ENCONTRADO EN COLUMNA E: '{str_value}'")
        return True
        
    except Exception:
        return False

def extract_client_data(client_value):
    """Extrae ID del cliente (dígitos antes del primer espacio) y razón social (resto)"""
    try:
        str_value = str(client_value).strip()
        # Buscar el primer espacio
        space_index = str_value.find(' ')
        if space_index > 0:
            id_cliente = str_value[:space_index]
            razon_social = str_value[space_index + 1:].strip()
            return id_cliente, razon_social
        return str_value, ""  # Si no hay espacio, todo es ID
    except Exception:
        return "", ""

def has_client_pattern(value):
    """Detecta si una celda tiene patrón de cliente (dígitos seguidos de espacio y texto)"""
    try:
        if value is None or pd.isna(value):
            return False
        
        str_value = str(value).strip()
        # Patrón: dígitos seguidos de espacio y más texto
        pattern = r'^\d+\s+.+'
        return bool(re.match(pattern, str_value))
        
    except Exception:
        return False

def process_csv_file(file_path):
    """Procesa archivos CSV con la nueva lógica especificada"""
    try:
        # Intentar leer el CSV con diferentes encodings
        encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding, header=None)
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            raise ValueError("No se pudo leer el archivo CSV con ninguna codificación soportada")
        
        print("PROCESANDO ARCHIVO CSV - LEYENDO TODOS LOS ELEMENTOS DE COLUMNA E...")
        
        processed_data = []
        current_row = 0
        max_rows = len(df)
        print(f"DEBUGGING: Total de filas en el archivo: {max_rows}")
        current_fecha = ""
        current_client_data = None
        
        while current_row < max_rows:
            # Obtener valores de las columnas principales
            col_a_value = df.iloc[current_row, 0] if len(df.columns) > 0 else None
            col_b_value = df.iloc[current_row, 1] if len(df.columns) > 1 else None
            col_e_value = df.iloc[current_row, 4] if len(df.columns) > 4 else None
            
            print(f"\n{'='*80}")
            print(f"PROCESANDO FILA {current_row + 1}")
            print(f"{'='*80}")
            print(f"Columna A: '{col_a_value}' (tipo: {type(col_a_value)})")
            print(f"Columna B: '{col_b_value}' (tipo: {type(col_b_value)})")
            print(f"Columna E: '{col_e_value}' (tipo: {type(col_e_value)})")
            
            # PRIMERO: SIEMPRE revisar si hay contenido en columna E (en cualquier tipo de línea)
            if has_content_in_column_e(col_e_value):
                print(f"🎯 *** CONTENIDO ENCONTRADO EN COLUMNA E - FILA {current_row + 1} ***")
                print(f"🎯 Valor raw de columna E: '{col_e_value}' (tipo: {type(col_e_value)})")
                
                fecha_procesada = clean_date_value(col_e_value)
                if fecha_procesada:
                    current_fecha = fecha_procesada
                    print(f"🎯 *** FECHA ASIGNADA EXITOSAMENTE: '{current_fecha}' ***")
                else:
                    print(f"🎯 *** ERROR: clean_date_value retornó vacío ***")
            
            # SEGUNDO: Determinar si es línea de tipo ARTÍCULO o NO ARTÍCULO
            if not is_empty_cell(col_a_value):
                print(f">>> TIPO: ARTÍCULO (columna A tiene contenido)")
                
                # Procesar como artículo
                id_articulo = clean_value_as_string(col_a_value)
                detalle_articulo = clean_value_as_string(df.iloc[current_row, 5] if len(df.columns) > 5 else '')
                cantidad_comprada = clean_numeric_value(df.iloc[current_row, 18] if len(df.columns) > 18 else '')
                precio_unitario = clean_numeric_value(df.iloc[current_row, 23] if len(df.columns) > 23 else '')
                descuento_1 = clean_numeric_value(df.iloc[current_row, 28] if len(df.columns) > 28 else '')
                descuento_2 = clean_numeric_value(df.iloc[current_row, 31] if len(df.columns) > 31 else '')
                descuento_3 = clean_numeric_value(df.iloc[current_row, 29] if len(df.columns) > 29 else '')
                total_con_descuentos = clean_numeric_value(df.iloc[current_row, 45] if len(df.columns) > 45 else '')
                
                if current_client_data is not None:
                    print(f">>> Usando fecha actual: '{current_fecha}'")
                    registro = {
                        'ID del Cliente': current_client_data['id_cliente'],
                        'Razon Social': current_client_data['razon_social'],
                        'Tipo de Documento': current_client_data['tipo_documento'],
                        'Serie del Documento': current_client_data['serie_documento'],
                        'ID del Documento': current_client_data['id_documento'],
                        'Fecha': current_fecha,
                        'Exento': current_client_data['exento'],
                        'Total Neto sin IVA': current_client_data['total_neto_sin_iva'],
                        'IVA Total del Documento': current_client_data['iva_total'],
                        'Total del Documento con IVA Incluido': current_client_data['total_con_iva'],
                        'Red': current_client_data['red'],
                        'ID de Articulo': id_articulo,
                        'Detalle de Articulo': detalle_articulo,
                        'Cantidad Comprada': cantidad_comprada,
                        'Precio Unitario': precio_unitario,
                        'Descuento 1 (%)': descuento_1,
                        'Descuento 2 (%)': descuento_2,
                        'Descuento 3 (%)': descuento_3,
                        'Total con Descuentos': total_con_descuentos
                    }
                    processed_data.append(registro)
                    print(f">>> ✅ Artículo procesado: ID={id_articulo} con fecha={current_fecha}")
                else:
                    print(f">>> ⚠️ Artículo encontrado pero sin cliente actual")
            
            else:
                print(f">>> TIPO: NO ARTÍCULO (columna A vacía)")
                
                # Verificar si es cliente
                if col_b_value is not None and has_client_pattern(col_b_value):
                    print(f">>> CLIENTE DETECTADO")
                    
                    id_cliente, razon_social = extract_client_data(col_b_value)
                    
                    tipo_documento = clean_value_as_string(df.iloc[current_row, 14] if len(df.columns) > 14 else '')
                    serie_documento = clean_value_as_string(df.iloc[current_row, 19] if len(df.columns) > 19 else '')
                    id_documento = clean_value_as_string(df.iloc[current_row, 20] if len(df.columns) > 20 else '')
                    exento = clean_numeric_value(df.iloc[current_row, 25] if len(df.columns) > 25 else '')
                    total_neto_sin_iva = clean_numeric_value(df.iloc[current_row, 29] if len(df.columns) > 29 else '')
                    iva_total = clean_numeric_value(df.iloc[current_row, 34] if len(df.columns) > 34 else '')
                    red = clean_numeric_value(df.iloc[current_row, 40] if len(df.columns) > 40 else '')
                    total_con_iva = clean_numeric_value(df.iloc[current_row, 44] if len(df.columns) > 44 else '')
                    
                    current_client_data = {
                        'id_cliente': id_cliente,
                        'razon_social': razon_social,
                        'tipo_documento': tipo_documento,
                        'serie_documento': serie_documento,
                        'id_documento': id_documento,
                        'exento': exento,
                        'total_neto_sin_iva': total_neto_sin_iva,
                        'iva_total': iva_total,
                        'red': red,
                        'total_con_iva': total_con_iva
                    }
                    
                    print(f">>> Cliente procesado: ID={id_cliente}, Razón={razon_social}")
                
                else:
                    print(f">>> Línea NO ARTÍCULO no identificada como cliente")
            
            print(f">>> Estado actual - Fecha: '{current_fecha}', Cliente: {current_client_data is not None}")
            current_row += 1
        
        print(f"\n{'='*80}")
        print(f"✅ PROCESAMIENTO COMPLETADO: {len(processed_data)} registros procesados")
        print(f"{'='*80}")
        
        if not processed_data:
            raise ValueError("No se encontraron datos válidos en el archivo")
        
        return processed_data
        
    except Exception as e:
        raise Exception(f"Error al procesar archivo CSV: {str(e)}")

def process_excel_file(file_path, file_extension):
    """Procesa archivos Excel (.xls y .xlsx) con la nueva lógica especificada"""
    try:
        if file_extension == '.xlsx':
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            max_row = sheet.max_row
            is_xlsx = True
        elif file_extension == '.xls':
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            max_row = sheet.nrows
            is_xlsx = False
        else:
            raise ValueError("Formato de archivo no soportado")

        processed_data = []

        def get_cell_value(row, col):
            try:
                if is_xlsx:
                    col_letter = get_column_letter(col)
                    cell = sheet[f'{col_letter}{row}']
                    return cell.value
                else:
                    return sheet.cell_value(row - 1, col - 1)
            except (IndexError, Exception):
                return None

        print("PROCESANDO ARCHIVO EXCEL - LEYENDO TODOS LOS ELEMENTOS DE COLUMNA E...")
        print(f"DEBUGGING: Total de filas en el archivo: {max_row}")
        
        current_row = 1
        current_fecha = ""
        current_client_data = None
        
        while current_row <= max_row:
            # Obtener valores de las columnas principales
            col_a_value = get_cell_value(current_row, 1)  # Columna A
            col_b_value = get_cell_value(current_row, 2)  # Columna B
            col_e_value = get_cell_value(current_row, 5)  # Columna E
            
            print(f"\n{'='*80}")
            print(f"PROCESANDO FILA {current_row}")
            print(f"{'='*80}")
            print(f"Columna A: '{col_a_value}' (tipo: {type(col_a_value)})")
            print(f"Columna B: '{col_b_value}' (tipo: {type(col_b_value)})")
            print(f"Columna E: '{col_e_value}' (tipo: {type(col_e_value)})")
            
            # PRIMERO: SIEMPRE revisar si hay contenido en columna E (en cualquier tipo de línea)
            if has_content_in_column_e(col_e_value):
                print(f"🎯 *** CONTENIDO ENCONTRADO EN COLUMNA E - FILA {current_row} ***")
                print(f"🎯 Valor raw de columna E: '{col_e_value}' (tipo: {type(col_e_value)})")
                
                fecha_procesada = clean_date_value(col_e_value)
                if fecha_procesada:
                    current_fecha = fecha_procesada
                    print(f"🎯 *** FECHA ASIGNADA EXITOSAMENTE: '{current_fecha}' ***")
                else:
                    print(f"🎯 *** ERROR: clean_date_value retornó vacío ***")
            
            # SEGUNDO: Determinar si es línea de tipo ARTÍCULO o NO ARTÍCULO
            if not is_empty_cell(col_a_value):
                print(f">>> TIPO: ARTÍCULO (columna A tiene contenido)")
                
                # Procesar como artículo
                id_articulo = clean_value_as_string(col_a_value)
                detalle_articulo = clean_value_as_string(get_cell_value(current_row, 6))
                cantidad_comprada = clean_numeric_value(get_cell_value(current_row, 19))
                precio_unitario = clean_numeric_value(get_cell_value(current_row, 24))
                descuento_1 = clean_numeric_value(get_cell_value(current_row, 29))
                descuento_2 = clean_numeric_value(get_cell_value(current_row, 32))
                descuento_3 = clean_numeric_value(get_cell_value(current_row, 30))
                total_con_descuentos = clean_numeric_value(get_cell_value(current_row, 46))
                
                if current_client_data is not None:
                    print(f">>> Usando fecha actual: '{current_fecha}'")
                    registro = {
                        'ID del Cliente': current_client_data['id_cliente'],
                        'Razon Social': current_client_data['razon_social'],
                        'Tipo de Documento': current_client_data['tipo_documento'],
                        'Serie del Documento': current_client_data['serie_documento'],
                        'ID del Documento': current_client_data['id_documento'],
                        'Fecha': current_fecha,
                        'Exento': current_client_data['exento'],
                        'Total Neto sin IVA': current_client_data['total_neto_sin_iva'],
                        'IVA Total del Documento': current_client_data['iva_total'],
                        'Total del Documento con IVA Incluido': current_client_data['total_con_iva'],
                        'Red': current_client_data['red'],
                        'ID de Articulo': id_articulo,
                        'Detalle de Articulo': detalle_articulo,
                        'Cantidad Comprada': cantidad_comprada,
                        'Precio Unitario': precio_unitario,
                        'Descuento 1 (%)': descuento_1,
                        'Descuento 2 (%)': descuento_2,
                        'Descuento 3 (%)': descuento_3,
                        'Total con Descuentos': total_con_descuentos
                    }
                    processed_data.append(registro)
                    print(f">>> ✅ Artículo procesado: ID={id_articulo} con fecha={current_fecha}")
                else:
                    print(f">>> ⚠️ Artículo encontrado pero sin cliente actual")
            
            else:
                print(f">>> TIPO: NO ARTÍCULO (columna A vacía)")
                
                # Verificar si es cliente
                if col_b_value is not None and has_client_pattern(col_b_value):
                    print(f">>> CLIENTE DETECTADO")
                    
                    id_cliente, razon_social = extract_client_data(col_b_value)
                    
                    tipo_documento = clean_value_as_string(get_cell_value(current_row, 15))
                    serie_documento = clean_value_as_string(get_cell_value(current_row, 20))
                    id_documento = clean_value_as_string(get_cell_value(current_row, 21))
                    exento = clean_numeric_value(get_cell_value(current_row, 26))
                    total_neto_sin_iva = clean_numeric_value(get_cell_value(current_row, 30))
                    iva_total = clean_numeric_value(get_cell_value(current_row, 35))
                    red = clean_numeric_value(get_cell_value(current_row, 41))
                    total_con_iva = clean_numeric_value(get_cell_value(current_row, 45))
                    
                    current_client_data = {
                        'id_cliente': id_cliente,
                        'razon_social': razon_social,
                        'tipo_documento': tipo_documento,
                        'serie_documento': serie_documento,
                        'id_documento': id_documento,
                        'exento': exento,
                        'total_neto_sin_iva': total_neto_sin_iva,
                        'iva_total': iva_total,
                        'red': red,
                        'total_con_iva': total_con_iva
                    }
                    
                    print(f">>> Cliente procesado: ID={id_cliente}, Razón={razon_social}")
                
                else:
                    print(f">>> Línea NO ARTÍCULO no identificada como cliente")
            
            print(f">>> Estado actual - Fecha: '{current_fecha}', Cliente: {current_client_data is not None}")
            current_row += 1

        print(f"\n{'='*80}")
        print(f"✅ PROCESAMIENTO COMPLETADO: {len(processed_data)} registros procesados")
        print(f"{'='*80}")

        if not processed_data:
            raise ValueError("No se encontraron datos válidos en el archivo")

        return processed_data

    except Exception as e:
        raise Exception(f"Error al procesar archivo Excel: {str(e)}")

def process_file(file_path, return_bytes=False):
    """
    Función principal que procesa archivos CSV, XLS y XLSX de ventas con la nueva lógica
    
    Args:
        file_path: Ruta del archivo a procesar
        return_bytes: Si True, devuelve bytes del archivo en lugar de guardarlo en disco
    
    Returns:
        Si return_bytes=True: tupla (bytes_data, filename)
        Si return_bytes=False: ruta del archivo guardado
    """
    try:
        file_extension = os.path.splitext(file_path)[1].lower()

        if file_extension == '.csv':
            processed_data = process_csv_file(file_path)
        elif file_extension in ['.xlsx', '.xls']:
            processed_data = process_excel_file(file_path, file_extension)
        else:
            raise ValueError("Formato de archivo no soportado. Solo se admiten .csv, .xlsx y .xls")

        if not processed_data:
            raise ValueError("No se encontraron datos válidos en el archivo")

        print(f"DEBUGGING: Total de registros procesados: {len(processed_data)}")
        if processed_data:
            print(f"DEBUGGING: Primer registro de ejemplo:")
            for key, value in processed_data[0].items():
                print(f"  {key}: '{value}'")

        # Crear DataFrame con los datos procesados
        df = pd.DataFrame(processed_data)

        # Asegurar el orden correcto de las columnas según especificación
        column_order = [
            'ID del Cliente',
            'Razon Social',
            'Tipo de Documento',
            'Serie del Documento',
            'ID del Documento',
            'Fecha',
            'Exento',
            'Total Neto sin IVA',
            'IVA Total del Documento',
            'Total del Documento con IVA Incluido',
            'Red',
            'ID de Articulo',
            'Detalle de Articulo',
            'Cantidad Comprada',
            'Precio Unitario',
            'Descuento 1 (%)',
            'Descuento 2 (%)',
            'Descuento 3 (%)',
            'Total con Descuentos'
        ]
        df = df[column_order]

        # Generar nombre del archivo de salida
        base_filename = os.path.splitext(os.path.basename(file_path))[0]
        output_filename = f"{base_filename}_PROCESADO.xlsx"

        if return_bytes:
            # Para aplicaciones web: devolver como bytes
            output_buffer = BytesIO()
            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Ventas Procesadas', index=False)
                
                # Ajustar el ancho de las columnas
                worksheet = writer.sheets['Ventas Procesadas']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output_buffer.seek(0)
            return output_buffer.getvalue(), output_filename
        else:
            # Para uso local: guardar archivo
            output_path = os.path.join(os.path.dirname(file_path), output_filename)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Ventas Procesadas', index=False)

                # Ajustar el ancho de las columnas
                worksheet = writer.sheets['Ventas Procesadas']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            return output_path

    except Exception as e:
        raise Exception(f"Error al procesar el archivo de ventas: {str(e)}")

def process_sales_data_for_webapp(file_bytes, original_filename):
    """
    Función específica para aplicaciones web que toma bytes del archivo
    
    Args:
        file_bytes: Bytes del archivo subido
        original_filename: Nombre del archivo original (se usa para generar el nombre de salida)
    
    Returns:
        Tupla (bytes_data, output_filename) del archivo procesado
    """
    import tempfile
    
    try:
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(original_filename)[1]) as temp_file:
            temp_file.write(file_bytes)
            temp_file_path = temp_file.name
        
        try:
            # Usar el nombre original del archivo para generar el nombre de salida
            base_filename = os.path.splitext(original_filename)[0]
            output_filename = f"{base_filename}_PROCESADO.xlsx"
            
            # Procesar el archivo temporal
            file_extension = os.path.splitext(original_filename)[1].lower()

            if file_extension == '.csv':
                processed_data = process_csv_file(temp_file_path)
            elif file_extension in ['.xlsx', '.xls']:
                processed_data = process_excel_file(temp_file_path, file_extension)
            else:
                raise ValueError("Formato de archivo no soportado. Solo se admiten .csv, .xlsx y .xls")

            if not processed_data:
                raise ValueError("No se encontraron datos válidos en el archivo")

            # Crear DataFrame con los datos procesados
            df = pd.DataFrame(processed_data)

            # Asegurar el orden correcto de las columnas
            column_order = [
                'ID del Cliente',
                'Razon Social',
                'Tipo de Documento',
                'Serie del Documento',
                'ID del Documento',
                'Fecha',
                'Exento',
                'Total Neto sin IVA',
                'IVA Total del Documento',
                'Total del Documento con IVA Incluido',
                'Red',
                'ID de Articulo',
                'Detalle de Articulo',
                'Cantidad Comprada',
                'Precio Unitario',
                'Descuento 1 (%)',
                'Descuento 2 (%)',
                'Descuento 3 (%)',
                'Total con Descuentos'
            ]
            df = df[column_order]

            # Generar archivo en memoria
            output_buffer = BytesIO()
            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Ventas Procesadas', index=False)
                
                # Ajustar el ancho de las columnas
                worksheet = writer.sheets['Ventas Procesadas']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output_buffer.seek(0)
            return output_buffer.getvalue(), output_filename
            
        finally:
            # Limpiar archivo temporal
            os.unlink(temp_file_path)
            
    except Exception as e:
        raise Exception(f"Error al procesar archivo para aplicación web: {str(e)}")

# Función de ejemplo para ejecutar el procesamiento
def main():
    """Función principal para probar el procesamiento"""
    try:
        # Solicitar la ruta del archivo
        file_path = input("Ingrese la ruta del archivo de ventas a procesar: ").strip()
        
        if not os.path.exists(file_path):
            print("Error: El archivo no existe.")
            return
        
        # Procesar el archivo
        output_path = process_file(file_path)
        print(f"\n¡Procesamiento completado exitosamente!")
        print(f"Archivo de salida guardado en: {output_path}")
        
    except Exception as e:
        print(f"Error durante el procesamiento: {str(e)}")

if __name__ == "__main__":
    main()

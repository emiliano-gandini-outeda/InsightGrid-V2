import pandas as pd
import openpyxl
import xlrd
import os
import csv
from datetime import datetime
from io import BytesIO
import re

def get_column_letter(col_num):
    """Genera letras de columna para Excel (A, B, C, ..., AA, AB, etc.)"""
    letters = []
    for i in range(1, 50):  # Ampliado para cubrir hasta AV (columna 48)
        if i <= 26:
            letters.append(chr(64 + i))
        else:
            first = chr(64 + ((i - 27) // 26) + 1)
            second = chr(64 + ((i - 27) % 26) + 1)
            letters.append(first + second)
    return letters[col_num - 1] if col_num <= len(letters) else f"Col{col_num}"

def clean_value_as_string(value, debug_info=""):
    """Convierte cualquier valor a string limpio, manteniendo el formato original"""
    try:
        if value is None:
            return ""
        elif pd.isna(value):
            return ""
        else:
            # Convertir a string y limpiar espacios
            str_value = str(value).strip()
            
            # Si es 'nan' como string, convertir a vacío
            if str_value.lower() in ['nan', 'none', 'null']:
                return ""
                
            return str_value
    except Exception:
        return ""

def clean_numeric_value_as_string(value):
    """Convierte un valor numérico a string, manteniendo formato pero limpiando"""
    try:
        if value is None or pd.isna(value):
            return "0"
        
        # Convertir a string y limpiar
        str_value = str(value).strip()
        
        # Si está vacío o es texto no numérico, devolver "0"
        if str_value == '' or str_value.lower() in ['nan', 'none', 'null']:
            return "0"
        
        # Intentar convertir a float para validar que es numérico, pero devolver como string
        try:
            float_val = float(str_value)
            # Si es un entero, devolverlo sin decimales
            if float_val.is_integer():
                return str(int(float_val))
            else:
                # Mantener hasta 2 decimales para valores monetarios
                return f"{float_val:.2f}"
        except (ValueError, TypeError):
            return "0"
            
    except (ValueError, TypeError):
        return "0"

def clean_id_value_as_string(value):
    """Convierte un ID a string, eliminando decimales innecesarios (.0)"""
    try:
        if value is None or pd.isna(value):
            return ""
        
        # Convertir a string y limpiar
        str_value = str(value).strip()
        
        # Si está vacío o es texto no válido, devolver vacío
        if str_value == '' or str_value.lower() in ['nan', 'none', 'null']:
            return ""
        
        # Si termina en .0, quitarlo
        if str_value.endswith('.0'):
            str_value = str_value[:-2]
        
        # Si es un número, intentar convertir y devolver sin decimales
        try:
            float_val = float(str_value)
            if float_val.is_integer():
                return str(int(float_val))
            else:
                # Para IDs, preferimos sin decimales, pero si tiene decimales los mantenemos
                return str_value
        except (ValueError, TypeError):
            # Si no es numérico, devolver tal como está
            return str_value
            
    except Exception:
        return ""

def is_empty_cell(value):
    """Verifica si una celda está vacía de manera más estricta y clara"""
    # Casos claramente vacíos
    if value is None:
        return True
        
    if pd.isna(value):
        return True
    
    # Convertir a string y verificar
    try:
        str_value = str(value).strip()
        
        # Casos de string vacío
        if str_value == '' or str_value.lower() in ['nan', 'none', 'null']:
            return True
            
        # Si llegamos aquí, la celda tiene contenido
        return False
        
    except Exception:
        return True

def is_vendedor_row(value):
    """Detecta si una celda contiene 'Vendedor'"""
    try:
        if value is None or pd.isna(value):
            return False
        
        # Convertir a string y limpiar espacios
        str_value = str(value).strip()
        
        # Buscar exactamente "Vendedor" (puede tener variaciones de espacios)
        return "Vendedor" in str_value
        
    except Exception:
        return False

def is_total_vendedor_row(value):
    """Detecta si una celda contiene 'Total Vendedor'"""
    try:
        if value is None or pd.isna(value):
            return False
        
        # Convertir a string y limpiar espacios
        str_value = str(value).strip()
        
        # Buscar exactamente "Total Vendedor"
        return "Total Vendedor" in str_value
        
    except Exception:
        return False

def process_csv_file(file_path):
    """Procesa archivos CSV buscando vendedores y sus artículos"""
    try:
        # Intentar leer el CSV con diferentes encodings
        encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding, header=None)
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            raise ValueError("No se pudo leer el archivo CSV con ninguna codificación soportada")
        
        print("PASO 1: Procesando vendedores y artículos...")
        processed_data = []
        
        current_row = 0
        vendedores_encontrados = 0
        current_vendedor_data = None
        
        while current_row < len(df):
            # VERIFICAR SI LA COLUMNA A contiene "Vendedor"
            col_a_value = df.iloc[current_row, 0] if len(df.columns) > 0 else None
            
            if is_vendedor_row(col_a_value):
                # ENCONTRÓ "Vendedor" - extraer datos del vendedor
                vendedores_encontrados += 1
                
                # Extraer ID del vendedor de columna B
                id_vendedor = clean_id_value_as_string(df.iloc[current_row, 1] if len(df.columns) > 1 else '')
                
                # Extraer nombre del vendedor de columna C
                nombre_vendedor = clean_value_as_string(df.iloc[current_row, 2] if len(df.columns) > 2 else '')
                
                current_vendedor_data = {
                    'id_vendedor': id_vendedor,
                    'nombre_vendedor': nombre_vendedor
                }
                
                print(f"Vendedor encontrado en fila {current_row + 1}: ID={id_vendedor}, Nombre={nombre_vendedor}")
                
                current_row += 1
                
            elif is_total_vendedor_row(col_a_value):
                # ENCONTRÓ "Total Vendedor" - saltar esta línea y la siguiente
                print(f"Total Vendedor encontrado en fila {current_row + 1}, saltando...")
                current_row += 2  # Saltar esta línea y la siguiente
                current_vendedor_data = None  # Resetear vendedor actual
                
            else:
                # LÍNEA DE ARTÍCULO (si hay un vendedor actual)
                if current_vendedor_data is not None:
                    # Extraer datos del artículo según las columnas especificadas
                    id_articulo = clean_id_value_as_string(df.iloc[current_row, 0] if len(df.columns) > 0 else '')  # Columna A
                    descripcion_articulo = clean_value_as_string(df.iloc[current_row, 1] if len(df.columns) > 1 else '')  # Columna B
                    cantidad = clean_numeric_value_as_string(df.iloc[current_row, 2] if len(df.columns) > 2 else '')  # Columna C
                    monto_s_iva = clean_numeric_value_as_string(df.iloc[current_row, 3] if len(df.columns) > 3 else '')  # Columna D
                    
                    # Solo procesar si hay datos válidos de artículo
                    if id_articulo.strip():
                        # Crear registro de artículo asociado al vendedor actual
                        articulo_data = {
                            'ID Vendedor': current_vendedor_data['id_vendedor'],
                            'Nombre Vendedor': current_vendedor_data['nombre_vendedor'],
                            'ID Articulo': id_articulo,
                            'Descripcion Articulo': descripcion_articulo,
                            'Cantidad': cantidad,
                            'Monto S-IVA': monto_s_iva
                        }
                        processed_data.append(articulo_data)
                        print(f"Artículo {id_articulo} en fila {current_row + 1} asignado a vendedor: {current_vendedor_data['nombre_vendedor']}")
                
                current_row += 1
        
        if not processed_data and vendedores_encontrados == 0:
            raise ValueError("No se encontraron vendedores válidos en el archivo. Verifique que el formato sea correcto.")
        
        return processed_data
        
    except Exception as e:
        raise Exception(f"Error al procesar archivo CSV: {str(e)}")

def process_excel_file(file_path, file_extension):
    """Procesa archivos Excel (.xls y .xlsx) buscando vendedores y sus artículos"""
    try:
        if file_extension == '.xlsx':
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            max_row = sheet.max_row
            max_col = sheet.max_column
            is_xlsx = True
        elif file_extension == '.xls':
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            max_row = sheet.nrows
            max_col = sheet.ncols
            is_xlsx = False
        else:
            raise ValueError("Formato de archivo no soportado")

        processed_data = []

        def get_cell_value(row, col):
            try:
                if is_xlsx:
                    col_letter = get_column_letter(col)
                    cell = sheet[f'{col_letter}{row}']
                    return cell.value
                else:
                    return sheet.cell_value(row - 1, col - 1)
            except (IndexError, Exception):
                return None

        print("PASO 1: Procesando vendedores y artículos...")
        current_row = 1
        vendedores_encontrados = 0
        current_vendedor_data = None

        while current_row <= max_row:
            # VERIFICAR SI LA COLUMNA A contiene "Vendedor"
            col_a_value = get_cell_value(current_row, 1)  # Columna A
            
            if is_vendedor_row(col_a_value):
                # ENCONTRÓ "Vendedor" - extraer datos del vendedor
                vendedores_encontrados += 1
                
                # Extraer ID del vendedor de columna B
                id_vendedor = clean_id_value_as_string(get_cell_value(current_row, 2))  # Columna B
                
                # Extraer nombre del vendedor de columna C
                nombre_vendedor = clean_value_as_string(get_cell_value(current_row, 3))  # Columna C
                
                current_vendedor_data = {
                    'id_vendedor': id_vendedor,
                    'nombre_vendedor': nombre_vendedor
                }
                
                print(f"Vendedor encontrado en fila {current_row}: ID={id_vendedor}, Nombre={nombre_vendedor}")
                
                current_row += 1
                
            elif is_total_vendedor_row(col_a_value):
                # ENCONTRÓ "Total Vendedor" - saltar esta línea y la siguiente
                print(f"Total Vendedor encontrado en fila {current_row}, saltando...")
                current_row += 2  # Saltar esta línea y la siguiente
                current_vendedor_data = None  # Resetear vendedor actual
                
            else:
                # LÍNEA DE ARTÍCULO (si hay un vendedor actual)
                if current_vendedor_data is not None:
                    # Extraer datos del artículo según las columnas especificadas
                    id_articulo = clean_id_value_as_string(get_cell_value(current_row, 1))  # Columna A
                    descripcion_articulo = clean_value_as_string(get_cell_value(current_row, 2))  # Columna B
                    cantidad = clean_numeric_value_as_string(get_cell_value(current_row, 3))  # Columna C
                    monto_s_iva = clean_numeric_value_as_string(get_cell_value(current_row, 4))  # Columna D
                    
                    # Solo procesar si hay datos válidos de artículo
                    if id_articulo.strip():
                        # Crear registro de artículo asociado al vendedor actual
                        articulo_data = {
                            'ID Vendedor': current_vendedor_data['id_vendedor'],
                            'Nombre Vendedor': current_vendedor_data['nombre_vendedor'],
                            'ID Articulo': id_articulo,
                            'Descripcion Articulo': descripcion_articulo,
                            'Cantidad': cantidad,
                            'Monto S-IVA': monto_s_iva
                        }
                        processed_data.append(articulo_data)
                        print(f"Artículo {id_articulo} en fila {current_row} asignado a vendedor: {current_vendedor_data['nombre_vendedor']}")

                current_row += 1

        if not processed_data and vendedores_encontrados == 0:
            raise ValueError("No se encontraron vendedores válidos en el archivo. Verifique que el formato sea correcto.")

        return processed_data

    except Exception as e:
        raise Exception(f"Error al procesar archivo Excel: {str(e)}")

def process_file(file_path, return_bytes=False):
    """
    Función principal que procesa archivos CSV, XLS y XLSX de vendedores
    
    Args:
        file_path: Ruta del archivo a procesar
        return_bytes: Si True, devuelve bytes del archivo en lugar de guardarlo en disco
    
    Returns:
        Si return_bytes=True: tupla (bytes_data, filename)
        Si return_bytes=False: ruta del archivo guardado
    """
    try:
        file_extension = os.path.splitext(file_path)[1].lower()

        if file_extension == '.csv':
            processed_data = process_csv_file(file_path)
        elif file_extension in ['.xlsx', '.xls']:
            processed_data = process_excel_file(file_path, file_extension)
        else:
            raise ValueError("Formato de archivo no soportado. Solo se admiten .csv, .xlsx y .xls")

        if not processed_data:
            raise ValueError("No se encontraron datos válidos en el archivo")

        print(f"DEBUGGING: Total de registros procesados: {len(processed_data)}")
        if processed_data:
            print(f"DEBUGGING: Primer registro de ejemplo:")
            for key, value in processed_data[0].items():
                print(f"  {key}: '{value}'")

        # Crear DataFrame con los datos procesados
        # IMPORTANTE: dtype=str para asegurar que todas las columnas sean strings
        df = pd.DataFrame(processed_data, dtype=str)

        # Asegurar el orden correcto de las columnas
        column_order = [
            'ID Vendedor',
            'Nombre Vendedor',
            'ID Articulo',
            'Descripcion Articulo',
            'Cantidad',
            'Monto S-IVA'
        ]
        df = df[column_order]

        # Generar nombre del archivo de salida
        base_filename = os.path.splitext(os.path.basename(file_path))[0]
        output_filename = f"{base_filename}_PROCESADO.xlsx"

        if return_bytes:
            # Para aplicaciones web: devolver como bytes
            output_buffer = BytesIO()
            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                # IMPORTANTE: Exportar con string_data=True para mantener formato de string
                df.to_excel(writer, sheet_name='Vendedores Procesados', index=False)
                
                # Ajustar el ancho de las columnas
                worksheet = writer.sheets['Vendedores Procesados']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output_buffer.seek(0)
            return output_buffer.getvalue(), output_filename
        else:
            # Para uso local: guardar archivo
            output_path = os.path.join(os.path.dirname(file_path), output_filename)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Vendedores Procesados', index=False)

                # Ajustar el ancho de las columnas
                worksheet = writer.sheets['Vendedores Procesados']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            return output_path

    except Exception as e:
        raise Exception(f"Error al procesar el archivo de vendedores: {str(e)}")

def process_vendedores_data_for_webapp(file_bytes, original_filename):
    """
    Función específica para aplicaciones web que toma bytes del archivo
    
    Args:
        file_bytes: Bytes del archivo subido
        original_filename: Nombre del archivo original (se usa para generar el nombre de salida)
    
    Returns:
        Tupla (bytes_data, output_filename) del archivo procesado
    """
    import tempfile
    
    try:
        # Crear archivo temporal
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(original_filename)[1]) as temp_file:
            temp_file.write(file_bytes)
            temp_file_path = temp_file.name
        
        try:
            # CORRECCIÓN: Usar el nombre original del archivo para generar el nombre de salida
            base_filename = os.path.splitext(original_filename)[0]
            output_filename = f"{base_filename}_PROCESADO.xlsx"
            
            # Procesar el archivo temporal usando la lógica existente
            file_extension = os.path.splitext(original_filename)[1].lower()

            if file_extension == '.csv':
                processed_data = process_csv_file(temp_file_path)
            elif file_extension in ['.xlsx', '.xls']:
                processed_data = process_excel_file(temp_file_path, file_extension)
            else:
                raise ValueError("Formato de archivo no soportado. Solo se admiten .csv, .xlsx y .xls")

            if not processed_data:
                raise ValueError("No se encontraron datos válidos en el archivo")

            # Crear DataFrame con los datos procesados
            # IMPORTANTE: dtype=str para asegurar que todas las columnas sean strings
            df = pd.DataFrame(processed_data, dtype=str)

            # Asegurar el orden correcto de las columnas
            column_order = [
                'ID Vendedor',
                'Nombre Vendedor',
                'ID Articulo',
                'Descripcion Articulo',
                'Cantidad',
                'Monto S-IVA'
            ]
            df = df[column_order]

            # Generar archivo en memoria
            output_buffer = BytesIO()
            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Vendedores Procesados', index=False)
                
                # Ajustar el ancho de las columnas
                worksheet = writer.sheets['Vendedores Procesados']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output_buffer.seek(0)
            return output_buffer.getvalue(), output_filename
            
        finally:
            # Limpiar archivo temporal
            os.unlink(temp_file_path)
            
    except Exception as e:
        raise Exception(f"Error al procesar archivo para aplicación web: {str(e)}")

# Función de ejemplo para ejecutar el procesamiento
def main():
    """Función principal para probar el procesamiento"""
    try:
        # Solicitar la ruta del archivo
        file_path = input("Ingrese la ruta del archivo de vendedores a procesar: ").strip()
        
        if not os.path.exists(file_path):
            print("Error: El archivo no existe.")
            return
        
        # Procesar el archivo
        output_path = process_file(file_path)
        print(f"\n¡Procesamiento completado exitosamente!")
        print(f"Archivo de salida guardado en: {output_path}")
        
    except Exception as e:
        print(f"Error durante el procesamiento: {str(e)}")

if __name__ == "__main__":
    main()

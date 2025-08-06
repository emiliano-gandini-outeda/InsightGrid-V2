import pandas as pd
import openpyxl
import xlrd
import os
import csv
from datetime import datetime

def get_column_letter(col_num):
    """Genera letras de columna para Excel (A, B, C, ..., AA, AB, etc.)"""
    letters = []
    for i in range(1, 35):
        if i <= 26:
            letters.append(chr(64 + i))
        else:
            letters.append('A' + chr(64 + i - 26))
    return letters[col_num - 1] if col_num <= len(letters) else f"Col{col_num}"

def clean_value(value, data_type='string'):
    """Limpia y convierte valores según el tipo especificado"""
    if pd.isna(value) or value == '' or str(value).strip() == '{Sin Definir}' or str(value).strip() == '':
        return "Dato no Definido"

    if data_type == 'integer':
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace(' ', '')
            return int(float(value))
        except (ValueError, TypeError):
            return "Dato no Definido"
    elif data_type == 'float':
        try:
            if isinstance(value, str):
                value = value.replace(',', '').replace(' ', '')
            return float(value)
        except (ValueError, TypeError):
            return "Dato no Definido"
    elif data_type == 'importado':
        str_value = str(value).strip().lower()
        if str_value in ['si', 'sí', 'yes', 'y', '1', 'true', 'verdadero']:
            return "Si"
        elif str_value in ['no', 'n', '0', 'false', 'falso']:
            return "No"
        else:
            return "No"
    else:
        return str(value).strip() if str(value).strip() != '' else "Dato no Definido"

def process_csv_file(file_path):
    """Procesa archivos CSV - asume que los datos están en un formato tabular simple"""
    try:
        # Intentar leer el CSV con diferentes encodings
        encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            raise ValueError("No se pudo leer el archivo CSV con ninguna codificación soportada")
        
        # Si el CSV tiene el formato esperado con proveedores, procesarlo
        # Si no, intentar mapear columnas conocidas
        processed_data = []
        
        # Buscar columnas que podrían contener la información necesaria
        columns = df.columns.tolist()
        
        # Mapeo de posibles nombres de columnas
        column_mapping = {
            'id_proveedor': ['id_proveedor', 'proveedor_id', 'id proveedor', 'supplier_id'],
            'nombre_proveedor': ['nombre_proveedor', 'proveedor', 'nombre proveedor', 'supplier_name', 'proveedor_nombre'],
            'id_articulo': ['id_articulo', 'articulo_id', 'id articulo', 'product_id', 'codigo'],
            'nombre_articulo': ['nombre_articulo', 'articulo', 'nombre articulo', 'product_name', 'descripcion'],
            'stock_minimo': ['stock_minimo', 'stock minimo', 'minimum_stock', 'min_stock'],
            'estado_producto': ['estado_producto', 'estado', 'status', 'estado producto'],
            'importado': ['importado', 'import', 'imported'],
            'codigo_proveedor': ['codigo_proveedor', 'codigo proveedor', 'supplier_code', 'codigo_supplier']
        }
        
        # Encontrar las columnas correspondientes
        found_columns = {}
        for key, possible_names in column_mapping.items():
            for col in columns:
                if col.lower().strip() in [name.lower() for name in possible_names]:
                    found_columns[key] = col
                    break
        
        # Si encontramos al menos algunas columnas clave, procesamos el archivo
        if len(found_columns) >= 2:  # Al menos 2 columnas identificadas
            for _, row in df.iterrows():
                articulo_data = {
                    'ID Proveedor': clean_value(row.get(found_columns.get('id_proveedor', ''), ''), 'integer'),
                    'Nombre Proveedor': clean_value(row.get(found_columns.get('nombre_proveedor', ''), ''), 'string'),
                    'ID Articulo': clean_value(row.get(found_columns.get('id_articulo', ''), ''), 'string'),
                    'Nombre Articulo': clean_value(row.get(found_columns.get('nombre_articulo', ''), ''), 'string'),
                    'Stock Minimo': clean_value(row.get(found_columns.get('stock_minimo', ''), ''), 'float'),
                    'Estado del Producto': clean_value(row.get(found_columns.get('estado_producto', ''), ''), 'string'),
                    'Importado': clean_value(row.get(found_columns.get('importado', ''), ''), 'importado'),
                    'Codigo para Proveedor': clean_value(row.get(found_columns.get('codigo_proveedor', ''), ''), 'string')
                }
                processed_data.append(articulo_data)
        else:
            # Si no podemos mapear automáticamente, tomar las primeras columnas disponibles
            for _, row in df.iterrows():
                row_values = row.tolist()
                articulo_data = {
                    'ID Proveedor': clean_value(row_values[0] if len(row_values) > 0 else '', 'integer'),
                    'Nombre Proveedor': clean_value(row_values[1] if len(row_values) > 1 else '', 'string'),
                    'ID Articulo': clean_value(row_values[2] if len(row_values) > 2 else '', 'string'),
                    'Nombre Articulo': clean_value(row_values[3] if len(row_values) > 3 else '', 'string'),
                    'Stock Minimo': clean_value(row_values[4] if len(row_values) > 4 else '', 'float'),
                    'Estado del Producto': clean_value(row_values[5] if len(row_values) > 5 else '', 'string'),
                    'Importado': clean_value(row_values[6] if len(row_values) > 6 else '', 'importado'),
                    'Codigo para Proveedor': clean_value(row_values[7] if len(row_values) > 7 else '', 'string')
                }
                processed_data.append(articulo_data)
        
        return processed_data
        
    except Exception as e:
        raise Exception(f"Error al procesar archivo CSV: {str(e)}")

def process_excel_file(file_path, file_extension):
    """Procesa archivos Excel (.xls y .xlsx) con el formato específico de proveedores"""
    try:
        if file_extension == '.xlsx':
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            max_row = sheet.max_row
            is_xlsx = True
        elif file_extension == '.xls':
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            max_row = sheet.nrows
            is_xlsx = False
        else:
            raise ValueError("Formato de archivo no soportado")

        processed_data = []

        def get_cell_value(row, col):
            if is_xlsx:
                col_letter = get_column_letter(col)
                cell = sheet[f'{col_letter}{row}']
                return cell.value
            else:
                try:
                    return sheet.cell_value(row - 1, col - 1)
                except IndexError:
                    return None

        current_row = 1
        proveedores_encontrados = 0

        while current_row <= max_row:
            cell_b_value = get_cell_value(current_row, 2)

            if cell_b_value and 'Proveedor:' in str(cell_b_value):
                proveedores_encontrados += 1
                id_proveedor = clean_value(get_cell_value(current_row, 6), 'integer')
                nombre_proveedor = clean_value(get_cell_value(current_row, 13), 'string')

                current_row += 1

                while current_row <= max_row:
                    cell_b_next_value = get_cell_value(current_row, 2)
                    if cell_b_next_value and 'Proveedor:' in str(cell_b_next_value):
                        break

                    id_articulo = clean_value(get_cell_value(current_row, 2), 'string')
                    nombre_articulo = clean_value(get_cell_value(current_row, 9), 'string')
                    stock_minimo = clean_value(get_cell_value(current_row, 19), 'float')
                    estado_producto = clean_value(get_cell_value(current_row, 22), 'string')
                    importado = clean_value(get_cell_value(current_row, 26), 'importado')
                    codigo_proveedor = clean_value(get_cell_value(current_row, 29), 'string')

                    if (id_articulo != "Dato no Definido" or nombre_articulo != "Dato no Definido"):
                        articulo_data = {
                            'ID Proveedor': id_proveedor,
                            'Nombre Proveedor': nombre_proveedor,
                            'ID Articulo': id_articulo,
                            'Nombre Articulo': nombre_articulo,
                            'Stock Minimo': stock_minimo,
                            'Estado del Producto': estado_producto,
                            'Importado': importado,
                            'Codigo para Proveedor': codigo_proveedor
                        }
                        processed_data.append(articulo_data)

                    current_row += 1
            else:
                current_row += 1

        if not processed_data and proveedores_encontrados == 0:
            # Si no encontramos el formato de proveedores, intentar leer como tabla normal
            try:
                df = pd.read_excel(file_path)
                processed_data = []
                
                for _, row in df.iterrows():
                    row_values = row.tolist()
                    articulo_data = {
                        'ID Proveedor': clean_value(row_values[0] if len(row_values) > 0 else '', 'integer'),
                        'Nombre Proveedor': clean_value(row_values[1] if len(row_values) > 1 else '', 'string'),
                        'ID Articulo': clean_value(row_values[2] if len(row_values) > 2 else '', 'string'),
                        'Nombre Articulo': clean_value(row_values[3] if len(row_values) > 3 else '', 'string'),
                        'Stock Minimo': clean_value(row_values[4] if len(row_values) > 4 else '', 'float'),
                        'Estado del Producto': clean_value(row_values[5] if len(row_values) > 5 else '', 'string'),
                        'Importado': clean_value(row_values[6] if len(row_values) > 6 else '', 'importado'),
                        'Codigo para Proveedor': clean_value(row_values[7] if len(row_values) > 7 else '', 'string')
                    }
                    processed_data.append(articulo_data)
            except:
                raise ValueError("No se encontraron datos válidos en el formato esperado")

        return processed_data

    except Exception as e:
        raise Exception(f"Error al procesar archivo Excel: {str(e)}")

def process_file(file_path):
    """Función principal que procesa archivos CSV, XLS y XLSX"""
    try:
        file_extension = os.path.splitext(file_path)[1].lower()

        if file_extension == '.csv':
            processed_data = process_csv_file(file_path)
        elif file_extension in ['.xlsx', '.xls']:
            processed_data = process_excel_file(file_path, file_extension)
        else:
            raise ValueError("Formato de archivo no soportado. Solo se admiten .csv, .xlsx y .xls")

        if not processed_data:
            raise ValueError("No se encontraron datos válidos en el archivo")

        # Crear DataFrame con los datos procesados
        df = pd.DataFrame(processed_data)

        # Asegurar el orden correcto de las columnas
        column_order = [
            'ID Proveedor',
            'Nombre Proveedor',
            'ID Articulo',
            'Nombre Articulo',
            'Stock Minimo',
            'Estado del Producto',
            'Importado',
            'Codigo para Proveedor'
        ]
        df = df[column_order]

        # Generar ruta de salida en el mismo directorio del archivo original
        original_dir = os.path.dirname(file_path)
        original_name = os.path.splitext(os.path.basename(file_path))[0]
        output_filename = f"{original_name}_PROCESADO.xlsx"
        output_path = os.path.join(original_dir, output_filename)

        # Guardar el archivo procesado
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Inventario Procesado', index=False)

            # Ajustar el ancho de las columnas
            worksheet = writer.sheets['Inventario Procesado']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        return output_path

    except Exception as e:
        raise Exception(f"Error al procesar el archivo de inventario: {str(e)}")